import math
import sqlite3
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, ttk
import numpy as np
import pandas as pd
import os
import locale
from openpyxl.cell.cell import MergedCell

import mysql.connector
import pandas as pd
from mysql.connector import Error
from openpyxl import load_workbook
from tkcalendar import DateEntry
from tkinterdnd2 import DND_FILES, TkinterDnD
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from decimal import Decimal

contacts = []
dataList1 = []
dataList2 = []
unique_values = set()
unique_values_sheet1 = set()

class Application(TkinterDnD.Tk):
    
    def __init__(self):
        
        super().__init__()
        Blank_space_title = " "
        self.title(100*Blank_space_title+"ระบบการจัดการรถหกล้อ")
        self.geometry("1440x900")

# สร้าง Notebook สำหรับเก็บแท็บ
        self.main_frame = ttk.Notebook(self)

# สร้างเฟรมต่างๆ
        self.frameAdd = SearchPage(parent=self.main_frame)
        self.frameSalary = FrameSalary(parent=self.main_frame)
        self.frameTax = FrameTax(parent=self.main_frame)
        self.frameSale = FrameSale(parent=self.main_frame)

# เพิ่มแต่ละเฟรมเข้าไปใน Notebook
        self.main_frame.add(self.frameSale, text='ตั้งเบิก')
        self.main_frame.add(self.frameAdd, text='เพิ่มข้อมูล')
        self.main_frame.add(self.frameSalary, text='คิดเงิน')
        self.main_frame.add(self.frameTax, text='ภาษี')

# ปรับขนาดและการขยายของ Notebook
        self.main_frame.pack(fill="both", expand=True) 

class DataTable(ttk.Treeview):
    def __init__(self, parent):
        super().__init__(parent)
        scroll_Y = tk.Scrollbar(self, orient="vertical", command=self.yview)
        scroll_X = tk.Scrollbar(self, orient="horizontal", command=self.xview)
        self.configure(yscrollcommand=scroll_Y.set, xscrollcommand=scroll_X.set)
        scroll_Y.pack(side="right", fill="y")
        scroll_X.pack(side="bottom", fill="x")
        self.stored_dataFrame = pd.DataFrame()

    def set_dataTable(self, dataFrame):
        self.stored_dataFrame = dataFrame
        self._draw_table(dataFrame) 

    def _draw_table(self, dataFrame):
        for item in self.get_children():
            self.delete(item)

        columns = list(dataFrame.columns)
        self["columns"] = columns 
        self["show"] = "headings" 

        for col in columns:
            self.heading(col, text=col)  
            self.column(col, width=100, anchor="center") 

        df_rows = dataFrame.to_numpy().tolist()
        for row in df_rows:
            self.insert("", "end", values=row)
    

class SearchPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        # self.file_names_listbox = tk.Listbox(parent, selectmode=tk.SINGLE, background="white")
        # self.file_names_listbox.place(rely=0.03,relheight=1, relwidth=0.15)
        # self.file_names_listbox.drop_target_register(DND_FILES)
        # self.file_names_listbox.dnd_bind("<<Drop>>", self.drop_inside_list_box)
        # self.file_names_listbox.bind("<Double-1>", self._display_file)
        
        self.file_names_listbox = tk.Listbox(self, selectmode=tk.SINGLE, background="white")
        self.file_names_listbox.place(rely=0.03,relheight=1, relwidth=0.15)
        self.file_names_listbox.drop_target_register(DND_FILES)
        self.file_names_listbox.dnd_bind("<<Drop>>", self.drop_inside_list_box)
        
        self.read_button = tk.Button(self, text="อ่านไฟล์ทั้งหมด", command=self._display_file)
        self.read_button.place(rely=0.95, relwidth=0.15)

        self.conform_button = tk.Button(parent,text = "ยืนยัน",command=self.saveData, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.clear_button = tk.Button(parent,text = "ล้าง",command=self.clearData, font=("Helvetica", 10), foreground="white", background="#dc3545", activebackground="#0069d9", activeforeground="white")
        
        self.conform_button.place(rely=0.94,relx=0.9,width=100)
        self.clear_button.place(rely=0.94,relx=0.82,width=100)

        # Treeview
        self.data_table = DataTable(parent)
        self.data_table.place(rely=0.03, relx=0.15, relwidth=0.85, relheight=0.85)

        self.path_map = {}                                                                                                                                     
#วนวาด listBox
    def drop_inside_list_box(self, event):
        file_paths = self._parse_drop_files(event.data)
        current_listbox_items = set(self.file_names_listbox.get(0, "end"))
        for file_path in file_paths:
            if file_path.endswith(".xls"):
                path_object = Path(file_path)
                file_name = path_object.name
                if file_name not in current_listbox_items:
                    self.file_names_listbox.insert("end", file_name)
                    self.path_map[file_name] = file_path
                

    def _display_file(self):
        contacts = []
        file_names = self.file_names_listbox.get(0, tk.END)  # ดึงชื่อไฟล์ทั้งหมดจาก Listbox
        for file_name in file_names:
            path = self.path_map[file_name]
            df = pd.read_excel(path, sheet_name='ใบสรุปรายวัน')  # อ่านข้อมูลจากไฟล์ Excel ใน sheet ที่ต้องการ
            data = df.values.tolist()
            
            for x in range(0,len(data)):
                # print(type(data[x][0]))
                if isinstance(data[x][0], int):
                    contacts.append((data[x][0],data[x][1],(data[x][2]),data[x][3],data[x][4],data[x][5],data[x][6],data[x][7],round(data[x][8],2),data[x][12],data[x][13],data[x][14]))
        data = pd.DataFrame(contacts,columns=['no','invoice', 'date', 'number_SN', 'weight', 'freight', 'down_coat','control','shipping','vehicle_registration','car_size','destination'])
        self.data_table.set_dataTable(dataFrame=data)
        return data
    
    
#เก็บ path ของไฟล
    def _parse_drop_files(self, filename):
            size = len(filename)
            res = []  # list of file paths
            name = ""
            idx = 0
            inside_braces = False  # เพิ่ม flag เพื่อตรวจสอบว่าอยู่ใน {}
    
            while idx < size:
                if filename[idx] == "{":
                    inside_braces = True  # เริ่มเจอ {}
                    name = ""  # ล้างค่า name เพราะจะเก็บไฟล์ใหม่ที่อยู่ใน {}
                elif filename[idx] == "}":
                    inside_braces = False  # เจอ } ให้ปิด flag
                    res.append(name)  # เก็บชื่อไฟล์ใน {}
                    name = ""  # ล้างค่า name หลังเก็บแล้ว
                elif filename[idx] == " " and not inside_braces:
                    if name != "":
                        res.append(name)  # เจอช่องว่างและ name ไม่ว่าง เก็บชื่อไฟล์ที่เจอ
                        name = ""  # ล้างค่า name เพื่อเริ่มเก็บไฟล์ถัดไป
                else:
                    name += filename[idx]  # รวมอักษรเป็นชื่อไฟล์
        
                idx += 1
    
            if name != "":
                res.append(name)  # เก็บชื่อไฟล์สุดท้ายที่เหลืออยู่
    
            return res
    
    def saveData(self):
        file_names = self.file_names_listbox.get(0, tk.END)  # ดึงชื่อไฟล์ทั้งหมดจาก Listbox
        for file_name in file_names:
            path = self.path_map[file_name]
            df = pd.read_excel(path, sheet_name='ใบสรุปรายวัน')  # อ่านข้อมูลจากไฟล์ Excel ใน sheet ที่ต้องการ
            data = df.values.tolist()
            for x in range(0, len(data)):
                if isinstance((data[x][0]),int):
                    contacts.append((data[x][0], data[x][1], data[x][2], data[x][3], data[x][4], data[x][5], data[x][6], data[x][7], data[x][8], data[x][12], data[x][13], data[x][14]))
        sqlQuery = "INSERT INTO truck (invoice, date, number_SN, weight, freight, down_coat,control,shipping,vehicle_registration,car_size,destination) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        try:
            connection = mysql.connector.connect(
            host='localhost',
            database='salary',
            user='root',
            password='')
            if connection.is_connected():
                try:   
                    for i in enumerate(contacts):
                        payload = (contacts[i[0]][1], contacts[i[0]][2], contacts[i[0]][3], contacts[i[0]][4], contacts[i[0]][5], contacts[i[0]][6], (None) if math.isnan(
                        contacts[i[0]][7])  else contacts[i[0]][7], round(contacts[i[0]][8], 2), contacts[i[0]][9], contacts[i[0]][10], contacts[i[0]][11])
                        cursor = connection.cursor()
                
                        cursor.execute(sqlQuery, payload)
                        connection.commit()
                    messagebox.showinfo("Success", "บันทึกเรียบร้อย")
                except mysql.connector.IntegrityError:
                        messagebox.showerror("Error", "มีข้อมูลที่ซ้ำกัน")
            cursor.close()
            connection.close()          
        except Error as e:
            print('Error:', e)
    def clearData(self):
        self.file_names_listbox.delete(0, "end")
        self.path_map.clear()
        contacts.clear()
class FrameSalary(tk.Frame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.pack(fill='both', expand=True)
        self.create_widgets()

    def create_widgets(self):
        # สร้าง Combobox
        self.vehicle_label = tk.Label(self, text="ทะเบียนรถ:")
        self.vehicle_label.place(rely=0.03, relx=0.03)
        
        options = ["รถก๊าช", "รถน้ำมัน", "รถร่วม"]
        self.variable = tk.StringVar(self)
        self.vehicle_entry = ttk.Combobox(self, textvariable=self.variable, values=options)
        self.vehicle_entry.current(0)
        self.vehicle_entry.place(rely=0.03, relx=0.1)

        self.from_date_label = tk.Label(self, text="วันที่เริ่มต้น:")
        self.from_date_label.place(rely=0.03, relx=0.27)

        self.from_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.from_date_entry.place(rely=0.03, relx=0.35)

        self.to_date_label = tk.Label(self, text="วันที่สิ้นสุด:")
        self.to_date_label.place(rely=0.03, relx=0.5)

        self.to_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.to_date_entry.place(rely=0.03, relx=0.58)

        self.search_button = tk.Button(self, text="นำข้อมูลเข้า Excel", command=self.search_data, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.search_button.place(rely=0.02, relx=0.7)

        self.data_table = ttk.Treeview(self, columns=('วันที่', 'ทะเบียนรถ', 'ปลายทาง', 'น้ำหนัก', 'ค่าควบ', 'ค่าขนลง'), show='headings')
        self.data_table.place(rely=0.07, relx=0.03, relwidth=0.94, relheight=0.85)

        for col in self.data_table["columns"]:
            self.data_table.heading(col, text=col)
            self.data_table.column(col, anchor="center")
        
        self.count_label = tk.Label(self, text="จำนวนข้อมูลทั้งหมด: 0")
        self.count_label.place(rely=0.95, relx=0.03)

    def search_data(self):
        vehicle_Name = self.vehicle_entry.get()
        if vehicle_Name == 'รถก๊าช':
            vehicle_number = ['71-1180','71-1481','72-1534','72-1535','72-1642','71-1761','72-1802','72-2148','71-2471','72-5479','72-6542','72-7328','72-7746','72-8057','72-8947']
        elif vehicle_Name == "รถร่วม":
            vehicle_number = ['73-0649','71-0865','72-1135','72-1933','72-1641','71-2039','72-2147','72-2593','70-3642','72-4029','71-6157','72-6772','72-7092','70-9849','71-5402','72-8750','83-3542','72-5467']
        else:
            vehicle_number = ['72-1533','72-3375','72-3976','70-9743','72-8058','72-1566']
        
        from_date = self.from_date_entry.get()
        to_date = self.to_date_entry.get()

        fromDate = datetime.strptime(from_date, "%m/%d/%y").strftime("%Y-%m-%d")
        toDate = datetime.strptime(to_date, "%m/%d/%y").strftime("%Y-%m-%d")
        
        try:
            connection = mysql.connector.connect(
            host='localhost',
            database='salary',
            user='root',
            password=''
            )
            cursor = connection.cursor()
            placeholders = ', '.join(['%s'] * len(vehicle_number))
            
            query = f"""
            SELECT s1.date, s1.vehicle_registration, s1.destination, s1.freight, s1.weight, s1.down_coat
            FROM truck AS s1
            WHERE s1.date BETWEEN %s AND %s
            AND s1.vehicle_registration IN ({placeholders})
            ORDER BY s1.vehicle_registration, s1.date
            """
            cursor.execute(query, [fromDate, toDate] + vehicle_number)
            data = cursor.fetchall()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Error: {err}")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
        
        data_df = pd.DataFrame(data, columns=['date', 'vehicle_registration', 'destination', 'freight', 'weight', 'down_coat'])
        self.display_table(data_df)
        self.update_count_label(len(data_df))
        self.export_data(data, fromDate, toDate)
    
    def display_table(self, data_df):
        # ลบข้อมูลในตารางเดิม
        for i in self.data_table.get_children():
            self.data_table.delete(i)
        
        # เพิ่มข้อมูลใหม่ลงในตาราง
        for row in data_df.itertuples():
            self.data_table.insert("", "end", values=row[1:])
    
    def update_count_label(self, count):
        self.count_label.config(text=f"จำนวนข้อมูลทั้งหมด: {count}")

    def export_data(self, data, from_date, to_date):
        template_path = 'F:\\salary\\salaryTruck.xlsx'
    # ตรวจสอบว่าไฟล์เทมเพลตมีอยู่
        if not os.path.exists(template_path):
            messagebox.showerror("Error", "ไฟล์เทมเพลตไม่พบ!")
            return
        
        workbook = load_workbook(template_path)
        selected_option = self.variable.get()  # ตัวเลือกจาก combobox
    
    # ตรวจสอบว่า sheet มีอยู่แล้วหรือไม่
        if selected_option in workbook.sheetnames:
            sheet = workbook[selected_option]
        else:
            messagebox.showerror("Error", f"ไม่มี sheet สำหรับ '{selected_option}' ในไฟล์ Excel")
            return
        # หาหัวข้อใน sheet และแผนที่ column index
        headers = ['วันที่', 'ทะเบียนรถ', 'ปลายทาง', 'ค่าขนส่งบ/ต', 'น้ำหนัก', 'ลงมือ']
        header_row = 1
        header_indices = {}
        
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value in headers:
                header_indices[cell_value] = col

    # ถ้าไม่มีหัวข้อใน sheet ให้แจ้งเตือน
        if not header_indices:
            messagebox.showerror("Error", "ไม่พบหัวข้อที่กำหนดไว้ใน sheet!")
            return
        # ตั้งชื่อไฟล์ใหม่และวนลูปตามทะเบียนรถ
        export_folder = 'F:\\salary\\exported'
        
        if selected_option == 'รถร่วม':
            export_folder = os.path.join(export_folder, 'รถร่วม')
        elif selected_option == 'รถน้ำมัน':
            export_folder = os.path.join(export_folder, 'รถน้ำมัน')
        else:
            export_folder = os.path.join(export_folder, 'รถก๊าช')
        if not os.path.exists(export_folder):
            os.makedirs(export_folder)
    
    # สร้าง dictionary เก็บข้อมูลตามทะเบียนรถ
        data_by_vehicle = {}
        for record in data:
            vehicle_number = record[1]  # ทะเบียนรถ
            if vehicle_number not in data_by_vehicle:
                data_by_vehicle[vehicle_number] = []
            data_by_vehicle[vehicle_number].append(record)
    
    # บันทึกไฟล์แยกตามทะเบียนรถ
        for vehicle_number, vehicle_data in data_by_vehicle.items():
        # โหลดไฟล์ใหม่สำหรับแต่ละทะเบียนรถ
            new_workbook = load_workbook(template_path)
            sheet = new_workbook[selected_option]
        
        # หาหัวข้อใน sheet และแผนที่ column index
            start_row = 2  # แถวเริ่มต้นที่ 2

            for record in vehicle_data:
                formatted_date = record[0].strftime('%Y-%m-%d')
                row_data = {
                    'วันที่': formatted_date,
                    'ทะเบียนรถ': record[1],
                    'ปลายทาง': record[2],
                    'ค่าขนส่งบ/ต': record[3],
                    'น้ำหนัก': record[4],
                    'ลงมือ': record[5],
                }

                for header, value in row_data.items():
                    if header in header_indices:
                        col_idx = header_indices[header]
                        target_cell = sheet.cell(row=start_row, column=col_idx)

            # ตรวจสอบว่าไม่ใช่ MergedCell ก่อนเขียนค่า
                        if not isinstance(target_cell, MergedCell):
                            target_cell.value = value
                        else:
                            print(f"Cannot write to merged cell at row {start_row}, column {col_idx}")
                start_row += 1  # ขยับไปแถวถัดไป
        # ตั้งชื่อไฟล์ใหม่สำหรับทะเบียนรถนี้
            from_date_formatted = datetime.strptime(from_date, "%Y-%m-%d").day
            to_date_formatted = datetime.strptime(to_date, "%Y-%m-%d").strftime("%d")
            month_year = datetime.strptime(to_date, "%Y-%m-%d").strftime("%b %y")  # เดือน/ปี
            month_thai = {
                "Jan": "ม.ค.",
                "Feb": "ก.พ.",
                "Mar": "มี.ค.",
                "Apr": "เม.ย.",
                "May": "พ.ค.",
                "Jun": "มิ.ย.",
                "Jul": "ก.ค.",
                "Aug": "ส.ค.",
                "Sep": "ก.ย.",
                "Oct": "ต.ค.",
                "Nov": "พ.ย.",
                "Dec": "ธ.ค."
            }
            month_en = month_year.split()[0]  # ได้เดือนภาษาอังกฤษ
            year = month_year.split()[1]      # ได้ปี

            month_thai_str = month_thai[month_en]  # แปลงเป็นเดือนภาษาไทย
            month_year_thai = f"{month_thai_str} {year[-2:]}"
            vehicle_number = vehicle_number[3:]
            new_file_name = f"{vehicle_number} {from_date_formatted}-{to_date_formatted} {month_year_thai}.xlsx"
            new_file_path = os.path.join(export_folder, new_file_name)

        # บันทึกไฟล์ใหม่สำหรับทะเบียนรถนี้
            new_workbook.save(new_file_path)
        messagebox.showinfo("Success", "ข้อมูลสำหรับทะเบียนรถถูกบันทึกลงในไฟล์ทั้งหมดแล้ว")
    # เริ่มใส่ข้อมูลจากแถวถัดไปของ sheet
        start_row = sheet.max_row + 1

    def display_table(self, dataframe):
        self.data_table.delete(*self.data_table.get_children())
        for row in dataframe.itertuples(index=False):
            self.data_table.insert("", "end", values=row)

    def update_count_label(self, count):
        self.count_label.config(text=f"จำนวนข้อมูลทั้งหมด: {count}")  

class FrameTax(tk.Frame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.pack(fill='both', expand=True)
        self.create_widgets()
        
    def create_widgets(self):
        # สร้าง Combobox
        self.vehicle_label = tk.Label(self, text="หจก:")
        self.vehicle_label.place(rely=0.03, relx=0.03)
        
        options = ["อนัสธิวัฒน์", "อานาปาน"]
        self.variable = tk.StringVar(self)
        self.vehicle_entry = ttk.Combobox(self, textvariable=self.variable, values=options)
        self.vehicle_entry.current(0)
        self.vehicle_entry.place(rely=0.03, relx=0.1)

        self.from_date_label = tk.Label(self, text="วันที่เริ่มต้น:")
        self.from_date_label.place(rely=0.03, relx=0.27)

        self.from_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.from_date_entry.place(rely=0.03, relx=0.35)

        self.to_date_label = tk.Label(self, text="วันที่สิ้นสุด:")
        self.to_date_label.place(rely=0.03, relx=0.5)

        self.to_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.to_date_entry.place(rely=0.03, relx=0.58)

        self.search_button = tk.Button(self, text="นำข้อมูลเข้า Excel", command=self.search_data, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.search_button.place(rely=0.02, relx=0.7)

        self.data_table = ttk.Treeview(self, columns=('วันที่', 'ทะเบียนรถ', 'ยอดขนส่ง'), show='headings')
        self.data_table.place(rely=0.07, relx=0.03, relwidth=0.94, relheight=0.85)

        for col in self.data_table["columns"]:
            self.data_table.heading(col, text=col)
            self.data_table.column(col, anchor="center")
        
        self.count_label = tk.Label(self, text="จำนวนข้อมูลทั้งหมด: 0")
        self.count_label.place(rely=0.95, relx=0.03)
        
    def search_data(self):
        vehicle_Name = self.vehicle_entry.get()
        if vehicle_Name == 'อนัสธิวัฒน์':
            vehicle_number = ['70-3642','71-0865','71-1761','71-2039', '72-1135', '72-1533', '72-1535', '72-1642', '72-1802'
            , '72-2147', '72-2593', '72-3976', '72-4029', '72-5467', '72-6772', '72-7092', '72-8750', '73-0649', '72-1534', '72-5479']          
        else :
            vehicle_number = ['71-1481','72-8947','71-2471','72-3375','71-1180','72-2148','72-6542','72-8057','72-7328','72-7746','72-1641','70-9849','72-1933','72-8058']
        from_date = self.from_date_entry.get()
        to_date = self.to_date_entry.get()
        
        fromDate = datetime.strptime(from_date, "%m/%d/%y").strftime("%Y-%m-%d")
        toDate = datetime.strptime(to_date, "%m/%d/%y").strftime("%Y-%m-%d")
        
        try:
            connection = mysql.connector.connect(
            host='localhost',
            database='salary',
            user='root',
            password=''
            )
            cursor = connection.cursor()
            placeholders = ', '.join(['%s'] * len(vehicle_number))
            
            query = f"""            
            SELECT s1.date,
            s1.shipping,
            s1.vehicle_registration
            FROM truck AS s1
            WHERE s1.date BETWEEN %s AND %s
            AND s1.vehicle_registration IN ({placeholders})
            ORDER BY s1.vehicle_registration, s1.date ASC;
            """
            cursor.execute(query, [fromDate, toDate] + vehicle_number)
            data = cursor.fetchall()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Error: {err}")
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
        data_df = pd.DataFrame(data, columns=['date', 'vehicle_registration', 'shipping'])
        
        self.display_table(data_df)
        self.update_count_label(len(data_df))
        self.export_data(data, fromDate, toDate, vehicle_number)
    def display_table(self, data_df):
        # ลบข้อมูลในตารางเดิม
        for i in self.data_table.get_children():
            self.data_table.delete(i)
        
        # เพิ่มข้อมูลใหม่ลงในตาราง
        for row in data_df.itertuples():
            self.data_table.insert("", "end", values=row[1:])
    def update_count_label(self, count):
        self.count_label.config(text=f"จำนวนข้อมูลทั้งหมด: {count}")  
        
    def export_data(self, data, from_date, to_date, vehicle_number):
        template_path = 'F:\\salary\\tax.xlsx'
    
    # ตรวจสอบว่าไฟล์เทมเพลตมีอยู่
        if not os.path.exists(template_path):
            messagebox.showerror("Error", "ไฟล์เทมเพลตไม่พบ!")
            return
    
    # โหลดไฟล์ Excel
        workbook = load_workbook(template_path)
        selected_option = self.variable.get()
    
    # ตรวจสอบว่ามีชีตหรือไม่
        if selected_option in workbook.sheetnames:
            sheet = workbook[selected_option]
        else:
            messagebox.showerror("Error", f"ไม่มี sheet สำหรับ '{selected_option}' ในไฟล์ Excel")
            return
    
    # ดึงค่าหัวตาราง
        headers = [vehicle_number]
        header_row = 1
        header_indices = {}
    
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value in headers:
                header_indices[cell_value] = col
    
    # จัดกลุ่มข้อมูลตามทะเบียนรถ
        data_by_vehicle = {}
        for record in data:
            v_number = record[2]  # ทะเบียนรถ
            if v_number not in data_by_vehicle:
                data_by_vehicle[v_number] = []
        
        # ดึงเฉพาะค่าที่เป็นตัวเลข (Decimal)
            try:
                decimal_value = Decimal(str(record[1]))  # ค่าที่ต้องบันทึก
                data_by_vehicle[v_number].append(decimal_value)
            except:
                continue
    
    # แปลงข้อมูลเป็น DataFrame
        updated_df = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in data_by_vehicle.items()]))
    
    # บันทึกลงไฟล์ Excel
        new_file_path = f"F:\\salary\\ภาษี {from_date}-{to_date}.xlsx"
        with pd.ExcelWriter(new_file_path, engine='xlsxwriter') as writer:
            updated_df.to_excel(writer, sheet_name=f"{from_date}-{to_date}", index=False)

        messagebox.showinfo("Success", f"บันทึกไฟล์สำเร็จ: {new_file_path}")
        
        
        
class FrameSale(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.file_names_listbox = tk.Listbox(self, selectmode=tk.SINGLE, background="white")
        self.file_names_listbox.place(rely=0.03,relheight=1, relwidth=0.15)
        self.file_names_listbox.drop_target_register(DND_FILES)
        self.file_names_listbox.dnd_bind("<<Drop>>", self.drop_inside_list_box)
        
        self.read_button = tk.Button(self, text="อ่านไฟล์ทั้งหมด", command=self._display_files)
        self.read_button.place(rely=0.95, relwidth=0.15)
        
        self.conform_button = tk.Button(self, text="ยืนยัน", command=self.export_data, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.conform_button.place(rely=0.94,relx=0.9,width=100)

        # Treeview
        self.data_table = DataTable(self)
        self.data_table.place(rely=0.03, relx=0.15, relwidth=0.85, relheight=0.85)

        self.path_map = {}  
    
    def drop_inside_list_box(self, event):
        file_paths = self._parse_drop_files(event.data)
        current_listbox_items = set(self.file_names_listbox.get(0, "end"))
        for file_path in file_paths:
            if file_path.endswith(".xls") or file_path.endswith(".xlsx"):
                path_object = Path(file_path)
                file_name = path_object.name
                if file_name not in current_listbox_items:
                    self.file_names_listbox.insert("end", file_name)
                    self.path_map[file_name] = file_path
    def _parse_drop_files(self, filename):
        size = len(filename)
        res = []  # list of file paths
        name = ""
        idx = 0
        while idx < size:
            if filename[idx] == "{":
                j = idx + 1
                while filename[j] != "}":
                    name += filename[j]
                    j += 1
                res.append(name)
                name = ""
                idx = j
            elif filename[idx] == " " and name != "":
                res.append(name)
                name = ""
            elif filename[idx] != " ":
                name += filename[idx]
            idx += 1
        if name != "":
            res.append(name)
        return res
    
    def _display_files(self):
        file_names = self.file_names_listbox.get(0, tk.END)  # ดึงชื่อไฟล์ทั้งหมดจาก Listbox
        for file_name in file_names:
            path = self.path_map[file_name]
            self._read_file(path)
    def _read_file(self, path):
        sheets_to_check = ['Sheet1', 'หจก.อนัสธิวัฒน์', 'บจก.อานาปาน 45']  # รายชื่อแผ่นงานที่ต้องการตรวจสอบ
        for sheet in sheets_to_check:
            try:
                df = pd.read_excel(path, sheet_name=sheet)
                data = df.values.tolist()
                for x in range(len(data)):
                    if sheet == 'Sheet1':
                        new_data = (data[x][4], data[x][0], data[x][1],data[x][8])
                        if isinstance(data[x][1], int) and data[x][1] not in unique_values_sheet1:
                            unique_values_sheet1.add(data[x][1])  # เพิ่มค่าลงในเซต
                            dataList1.append(new_data)  # เพิ่มค่าลงใน dataList1
                    if sheet == 'Sheet1':
                        new_data = (data[x][4], data[x][0], data[x][1])
                        if isinstance(data[x][1], int) and data[x][1] not in unique_values_sheet1:  # ตรวจสอบว่าเป็นค่าที่ไม่ซ้ำ
                            unique_values_sheet1.add(data[x][1])  # เพิ่มค่าลงในเซตเพื่อไม่ให้ซ้ำ
                            dataList1.append(new_data)  # เพิ่มค่าลงใน dataList1 
                    elif sheet == 'หจก.อนัสธิวัฒน์':
                        if isinstance(data[x][14], int) and data[x][14] not in unique_values:  # ตรวจสอบว่าเป็น int และไม่ซ้ำ
                            unique_values.add(data[x][14])  # เพิ่มค่าลงในเซตเพื่อไม่ให้ซ้ำ
                            dataList2.append((data[x][5], data[x][6], data[x][8], data[x][9], data[x][10], data[x][11], data[x][12], data[x][14], data[x][15], data[x][16]))
                    else:  # สำหรับแผ่นงานอื่น ๆ เช่น 'บจก.อานาปาน 45'
                        if isinstance(data[x][14], int) and data[x][14] not in unique_values:  # ตรวจสอบว่าเป็น int และไม่ซ้ำ
                            unique_values.add(data[x][14])  # เพิ่มค่าลงในเซตเพื่อไม่ให้ซ้ำ
                            dataList2.append((data[x][5], data[x][6], data[x][8], data[x][9], data[x][10], data[x][11], data[x][12], data[x][14], data[x][15], data[x][16]))
            # ไม่มีการ break ที่นี่ เพื่อให้ทำการอ่านแผ่นงานต่อไป
            except ValueError:
                continue  # ถ้าพบ ValueError ก็ข้ามไปอ่านแผ่นงานถัดไป 
        combined_df = self.formatData(dataList1, dataList2) 
        self.data_table.set_dataTable(combined_df)

    # เรียกใช้ formatData หลังจากประมวลผลแผ่นงานทั้งหมดเสร็จ
    # แสดง DataFrame ที่รวมกัน หรือจัดการตามต้องการ
    def export_data(self):
        # Call the method and pass the combined DataFrame
        combined_df = self.formatData(dataList1, dataList2) 
        self.data_table.set_dataTable(combined_df)
        self.exportExcel(combined_df)    
    
    def formatData(self, dataList1, dataList2):
        license_list = [
        '72-4029', '72-3976', '72-2593', '72-1135', '71-2039',
        '72-5479', '71-0865', '72-1642', '72-6772', '72-7092',
        '72-8750', '72-1534', '72-1802', '70-3642', '71-1761'
        ]

        # แปลง dataList1 เป็น DataFrame
        # df1 = pd.DataFrame(dataList1, columns=['Column1', 'Column2', 'Column3'])  # ตั้งชื่อคอลัมน์ตามต้องการ
        df1 = pd.DataFrame(dataList1, columns=['Column1', 'Column2', 'Column3', 'Column4'])  # ตั้งชื่อคอลัมน์ตามต้องการ กรณีมีสอง หจก

        # แปลง dataList2 เป็น DataFrame
        df2 = pd.DataFrame(dataList2, columns=[
        'ColumnA', 'ColumnB', 'ColumnC', 'ColumnD', 'ColumnE', 
        'ColumnF', 'ColumnG', 'ColumnH', 'ColumnI', 'ColumnJ'
        ])  # ตั้งชื่อคอลัมน์ตามต้องการ

    # รวม DataFrame ทั้งสองเข้าด้วยกัน
        combined_df = pd.merge(df1, df2, left_on='Column3', right_on='ColumnH', how='outer')
        

    # ฟังก์ชันจัดรูปแบบ Column1
        def format_column1(value):
            if isinstance(value, str) and ',' in value:
                parts = value.split(',')
                formatted = parts[0] + ',' + ','.join([x[-2:] for x in parts[1:3]])
                return formatted
            return value

    # ใช้ฟังก์ชันจัดรูปแบบกับ Column1
        combined_df['Column1'] = combined_df['Column1'].apply(format_column1)

    # จัดรูปแบบ ColumnI
        combined_df['ColumnI'] = combined_df['ColumnI'].apply(lambda x: x[2:] if isinstance(x, str) else x)

    # เพิ่มคอลัมน์ 'ขนาดรถ'
        combined_df['ขนาดรถ'] = combined_df['ColumnI'].apply(lambda x: '5.5 M' if x in license_list else '5 M')

    # คำนวณผลรวมของ ColumnD, ColumnE, ColumnF, ColumnG
        combined_df['ค่าควบ'] = combined_df[['ColumnD', 'ColumnE', 'ColumnF', 'ColumnG']].sum(axis=1)

    # ลบคอลัมน์ ColumnD, ColumnE, ColumnF, ColumnG
        combined_df.drop(columns=['ColumnD', 'ColumnE', 'ColumnF', 'ColumnG'], inplace=True)

    # ย้ายคอลัมน์ Sum ไปยังตำแหน่งเดิมของ ColumnD
        columns = combined_df.columns.tolist()
        sum_index = columns.index('ColumnH')  # ตำแหน่งใหม่ที่ต้องการ
        columns.remove('ค่าควบ')
        columns.insert(sum_index, 'ค่าควบ')
        combined_df = combined_df[columns]
        
        columns = combined_df.columns.tolist()
        columns.remove('ขนาดรถ')  # เอา NewColumn ออกจากลิสต์ก่อน
        columns.insert(columns.index('ColumnJ'), 'ขนาดรถ')  # ใส่ NewColumn ก่อนหน้า ColumnJ
        combined_df = combined_df[columns]

        if 'ColumnH' in combined_df.columns:
            combined_df.drop(columns=['ColumnH'], inplace=True)

    # แปลง Column2 เป็น datetime
        combined_df['Column2'] = pd.to_datetime(combined_df['Column2'], errors='coerce')

    # เรียงข้อมูลตามวันที่ใน Column2 และ Column4
        combined_df.sort_values(by=['Column4', 'Column2'], inplace=True)

    # ส่งข้อมูลกลับไปยัง DataTable ใน GUI
        self.data_table.set_dataTable(combined_df)
        return combined_df  # ส่งคืน DataFrame ที่รวมกัน
        
    def exportExcel(self, dataFrame):
    # เปิดไฟล์ Excel ที่มีอยู่
        file_path = "fomatData.xlsx"  # กำหนดเส้นทางไฟล์ที่มีอยู่
        try:
        # โหลดไฟล์ Excel
            book = load_workbook(file_path)
        # ตรวจสอบว่าแผ่นงานแรกมีชื่อว่า "Sheet1" หรือไม่
            if "Sheet1" not in book.sheetnames:
            # ถ้าไม่มี ให้สร้างแผ่นงานใหม่
                sheet = book.create_sheet("Sheet1")
            else:
            # ถ้ามีให้ใช้แผ่นงานที่มีอยู่
                sheet = book["Sheet1"]
        # ลบข้อมูลเก่าในแผ่นงาน (ถ้าต้องการ)
            for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
        # เขียน DataFrame ลงในแผ่นงาน
            for r_idx, row in enumerate(dataFrame.itertuples(index=False), start=2):  # เริ่มต้นที่แถวที่ 2
                for c_idx, value in enumerate(row, start=1):  # เริ่มต้นที่คอลัมน์ที่ 1
                    sheet.cell(row=r_idx, column=c_idx, value=value)
        # บันทึกไฟล์
            book.save(file_path)
            messagebox.showinfo("Success", f"Data exported successfully to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {str(e)}") 
        

if __name__ == "__main__":
    root = Application()
    root.mainloop()