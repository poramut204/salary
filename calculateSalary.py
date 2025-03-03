import math
import sqlite3
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import messagebox, ttk

import mysql.connector
import pandas as pd
from mysql.connector import Error
from openpyxl import load_workbook
from tkcalendar import DateEntry
from tkinterdnd2 import DND_FILES, TkinterDnD

contacts = []

class Application(TkinterDnD.Tk):
    
    def __init__(self):
        
        super().__init__()
        Blank_space_title = " "
        self.title(100*Blank_space_title+"ระบบการจัดการรถหกล้อ")
        self.main_frame = ttk.Notebook(self)
        self.geometry("1440x900")
        self.frameAdd = SearchPage(parent=self.main_frame)
        self.frameSalary = FrameSalary(parent=self.main_frame)
        self.frameTax = FrameTax(parent=self.main_frame)
        self.main_frame.add(self.frameAdd, text='เพิ่มข้อมูล')
        self.main_frame.add(self.frameSalary, text='คิดเงิน')
        self.main_frame.add(self.frameTax, text='ภาษี')
        self.main_frame.pack(fill="both", expand="true")

class DataTable(ttk.Treeview):
    def __init__(self, parent):
        super().__init__(parent)
        scroll_Y = tk.Scrollbar(self, orient="vertical", command=self.yview)
        scroll_X = tk.Scrollbar(self, orient="horizontal", command=self.xview)
        self.configure(yscrollcommand=scroll_Y.set, xscrollcommand=scroll_X.set)
        scroll_Y.pack(side="right", fill="y")
        scroll_X.pack(side="bottom", fill="x")
        self.stored_dataFrame = pd.DataFrame()

    def set_dataTable(self, dataFrame):
        self.stored_dataFrame = dataFrame
        self._draw_table(dataFrame) 

    def _draw_table(self, dataFrame):
        for item in self.get_children():
            self.delete(item)

        columns = list(dataFrame.columns)
        self["columns"] = columns 
        self["show"] = "headings" 

        for col in columns:
            self.heading(col, text=col)  
            self.column(col, width=100, anchor="center") 

        df_rows = dataFrame.to_numpy().tolist()
        for row in df_rows:
            self.insert("", "end", values=row)
    

class SearchPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.file_names_listbox = tk.Listbox(parent, selectmode=tk.SINGLE, background="white")
        self.file_names_listbox.place(rely=0.03,relheight=1, relwidth=0.15)
        self.file_names_listbox.drop_target_register(DND_FILES)
        self.file_names_listbox.dnd_bind("<<Drop>>", self.drop_inside_list_box)
        self.file_names_listbox.bind("<Double-1>", self._display_file)

        self.conform_button = tk.Button(parent,text = "ยืนยัน",command=self.saveData, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.clear_button = tk.Button(parent,text = "ล้าง",command=self.clearData, font=("Helvetica", 10), foreground="white", background="#dc3545", activebackground="#0069d9", activeforeground="white")
        
        self.conform_button.place(rely=0.94,relx=0.9,width=100)
        self.clear_button.place(rely=0.94,relx=0.82,width=100)

        # Treeview
        self.data_table = DataTable(parent)
        self.data_table.place(rely=0.03, relx=0.15, relwidth=0.85, relheight=0.85)

        self.path_map = {}                                                                                                                                     
#วนวาด listBox
    def drop_inside_list_box(self, event):
        file_paths = self._parse_drop_files(event.data)
        current_listbox_items = set(self.file_names_listbox.get(0, "end"))
        for file_path in file_paths:
            if file_path.endswith(".xls"):
                path_object = Path(file_path)
                file_name = path_object.name
                if file_name not in current_listbox_items:
                    self.file_names_listbox.insert("end", file_name)
                    self.path_map[file_name] = file_path
                

    def _display_file(self, event):
        file_name = self.file_names_listbox.get(self.file_names_listbox.curselection())
        path = self.path_map[file_name]
        df = pd.read_excel(path,sheet_name ='ใบสรุปรายวัน')
        # data = pd.DataFrame(df, columns=['ลำดับที่','เลขที่ใบส่งของ', 'วันที่', 'เลข SN', 'น้ำหนัก', 'ค่าขนส่ง', 'ค่าขนลง', 'ค่าควบ', 'รวมค่าขนส่ง' ])
        data = df.values.tolist()
            
        for x in range(0,len(data)):
                # print(type(data[x][0]))
            if type(data[x][0]) == int:
                contacts.append((data[x][0],data[x][1],(data[x][2]),data[x][3],data[x][4],data[x][5],data[x][6],data[x][7],round(data[x][8],2),data[x][12],data[x][13],data[x][14]))
        data = pd.DataFrame(contacts,columns=['no','invoice', 'date', 'number_SN', 'weight', 'freight', 'down_coat','control','shipping','vehicle_registration','car_size','destination'])
        self.data_table.set_dataTable(dataFrame=data)
        return data
    
    
#เก็บ path ของไฟล
    def _parse_drop_files(self, filename):
        size = len(filename)
        res = []  # list of file paths
        name = ""
        idx = 0
        while idx < size:
            if filename[idx] == "{":
                j = idx + 1
                while filename[j] != "}":
                    name += filename[j]
                    j += 1
                res.append(name)
                name = ""
                idx = j
            elif filename[idx] == " " and name != "":
                res.append(name)
                name = ""
            elif filename[idx] != " ":
                name += filename[idx]
            idx += 1
        if name != "":
            res.append(name)
        return res
    
    def saveData(self):
        file_name = self.file_names_listbox.get(self.file_names_listbox.curselection())
        path = self.path_map[file_name]
        df = pd.read_excel(path, sheet_name='ใบสรุปรายวัน')
        data = df.values.tolist()
        
        contacts = []
        for x in range(0, len(data)):
            if type(data[x][0]) == int:
                contacts.append((data[x][0], data[x][1], data[x][2], data[x][3], data[x][4], data[x][5], data[x][6], data[x][7], data[x][8], data[x][12], data[x][13], data[x][14]))

        data = pd.DataFrame(contacts, columns=['no', 'invoice', 'date', 'number_SN', 'weight', 'freight', 'down_coat', 'control', 'shipping', 'vehicle_registration', 'car_size', 'destination'])
        self.data_table.set_dataTable(dataFrame=data)

        sqlQuery = "INSERT INTO truck (invoice, date, number_SN, weight, freight, down_coat,control,shipping,vehicle_registration,car_size,destination) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        try:
            connection = mysql.connector.connect(
            host='localhost',
            database='anuttiwattruck',
            user='root',
            password='')
            if connection.is_connected():
                try:   
                    for i in enumerate(contacts):
                        payload = (contacts[i[0]][1], contacts[i[0]][2], contacts[i[0]][3], contacts[i[0]][4], contacts[i[0]][5], contacts[i[0]][6], (None) if math.isnan(
                        contacts[i[0]][7])  else contacts[i[0]][7], round(contacts[i[0]][8], 2), contacts[i[0]][9], contacts[i[0]][10], contacts[i[0]][11])
                        cursor = connection.cursor()
                
                        cursor.execute(sqlQuery, payload)
                        connection.commit()
                    messagebox.showinfo("Success", "บันทึกเรียบร้อย")
                except mysql.connector.IntegrityError:
                        messagebox.showerror("Error", "มีข้อมูลที่ซ้ำกัน")
            
            cursor.close()
            connection.close()
                            
        except Error as e:
            print('Error:', e)

    def clearData(self):
        self.file_names_listbox.delete(0, "end")
        self.path_map.clear()
class FrameSalary(tk.Frame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.pack(fill='both', expand=True)
        self.create_widgets()

    def create_widgets(self):
        # สร้าง Combobox
        self.vehicle_label = tk.Label(self, text="ทะเบียนรถ:")
        self.vehicle_label.place(rely=0.03, relx=0.03)
        
        options = ["รถก๊าช", "รถน้ำมัน", "รถร่วม"]
        self.variable = tk.StringVar(self)
        self.vehicle_entry = ttk.Combobox(self, textvariable=self.variable, values=options)
        self.vehicle_entry.current(0)
        self.vehicle_entry.place(rely=0.03, relx=0.1)

        self.from_date_label = tk.Label(self, text="วันที่เริ่มต้น:")
        self.from_date_label.place(rely=0.03, relx=0.27)

        self.from_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.from_date_entry.place(rely=0.03, relx=0.35)

        self.to_date_label = tk.Label(self, text="วันที่สิ้นสุด:")
        self.to_date_label.place(rely=0.03, relx=0.5)

        self.to_date_entry = DateEntry(self, background="lightblue", fieldbackground="white")
        self.to_date_entry.place(rely=0.03, relx=0.58)

        self.search_button = tk.Button(self, text="นำข้อมูลเข้า Excel", command=self.search_data, font=("Helvetica", 10), foreground="white", background="#007bff", activebackground="#0069d9", activeforeground="white")
        self.search_button.place(rely=0.02, relx=0.7)

        self.data_table = ttk.Treeview(self, columns=('วันที่', 'ทะเบียนรถ', 'ปลายทาง', 'น้ำหนัก', 'ค่าควบ', 'ค่าขนลง'), show='headings')
        self.data_table.place(rely=0.07, relx=0.03, relwidth=0.94, relheight=0.85)

        for col in self.data_table["columns"]:
            self.data_table.heading(col, text=col)
            self.data_table.column(col, anchor="center")
        
        self.count_label = tk.Label(self, text="จำนวนข้อมูลทั้งหมด: 0")
        self.count_label.place(rely=0.95, relx=0.03)

    def search_data(self):
        vehicle_Name = self.vehicle_entry.get()
        if vehicle_Name == 'รถก๊าช':
            vehicle_number = ['71-1180','71-1481','72-1534', '72-1535','72-1642','71-1761','72-1802','72-2148','71-2471','72-5479', '72-6542','72-7328','72-7746','72-8057','72-8947']
        elif vehicle_Name == "รถร่วม":
            vehicle_number = ['73-0649','71-0865','72-1135', '72-1566','72-1933','72-1641','71-2039','72-2147','72-2593','70-3642', '72-4029','71-6157','72-6772','72-7092','70-9849','71-5402','72-8750','83-3542','72-5467']
        else:
            vehicle_number = ['72-1533','72-3375', '72-3976','70-9743','72-8058']
        
        from_date = self.from_date_entry.get()
        to_date = self.to_date_entry.get()

        fromDate = datetime.strptime(from_date, "%m/%d/%y").strftime("%Y-%m-%d")
        toDate = datetime.strptime(to_date, "%m/%d/%y").strftime("%Y-%m-%d")

        connection = mysql.connector.connect(
            host='localhost',
            database='anuttiwattruck',
            user='root',
            password=''
        )
        cursor = connection.cursor()
    
        placeholders = ', '.join(['%s'] * len(vehicle_number))
    
        query = f"""
            SELECT s1.date, s1.vehicle_registration, s1.destination, s1.freight, s1.weight, s1.down_coat
            FROM truck AS s1
            WHERE s1.date BETWEEN %s AND %s
            AND s1.vehicle_registration IN ({placeholders})
            ORDER BY s1.vehicle_registration, s1.date
            """
        cursor.execute(query, [fromDate, toDate] + vehicle_number)
        data = cursor.fetchall()

        cursor.close()
        connection.close()
        
        data_df = pd.DataFrame(data, columns=['date', 'vehicle_registration', 'destination', 'freight', 'weight', 'down_coat'])
        self.display_table(data_df)
        self.update_count_label(len(data_df))
        self.export_data(data)
    
    def export_data(self, data) :
        data_by_registration = {}
        for record in data:
            formatted_date = record[0].strftime('%Y-%m-%d')
            registration = record[1]
            formatted_record = (formatted_date, *record[1:])
            # เก็บข้อมูลเฉพาะค่าเดียวของ Registration
            if registration not in data_by_registration:
                data_by_registration[registration] = []
            data_by_registration[registration].append(formatted_record)

        # สร้าง ExcelWriter object เพื่อจัดการการเขียนหลาย sheet
        file_path = 'F:\salary\data_by_registration.xlsx'

        # เพิ่มข้อมูลใหม่ลงใน workbook
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for registration, records in data_by_registration.items():
                df = pd.DataFrame(records, columns=['Date', 'Registration', 'Route', 'Distance', 'Fuel', 'Cost'])
        
                # เปลี่ยนชื่อคอลัมน์ Date เป็น 'วันที่'
                df.columns = ['วันที่', 'Registration', 'Route', 'Distance', 'Fuel', 'Cost']
        
            # ตรวจสอบว่า sheet มีอยู่แล้วหรือไม่
                if registration in writer.book.sheetnames:
                    start_row = writer.book[registration].max_row + 1
                    df.to_excel(writer, sheet_name=registration, index=False, header=False, startrow=start_row)
                else:
                    df.to_excel(writer, sheet_name=registration, index=False)
        messagebox.showinfo("Success", "ข้อมูลถูกเพิ่มลงในไฟล์ Excel เรียบร้อยแล้ว")

    def display_table(self, dataframe):
        self.data_table.delete(*self.data_table.get_children())
        for row in dataframe.itertuples(index=False):
            self.data_table.insert("", "end", values=row)
    def update_count_label(self, count):
        self.count_label.config(text=f"จำนวนข้อมูลทั้งหมด: {count}")  
    

class FrameTax(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.label = tk.Label(self, text="This is Tax Page")
        self.label.pack()
        self.button = tk.Button(self, text="Click Me")
        self.button.pack()

if __name__ == "__main__":
    root = Application()
    root.mainloop()